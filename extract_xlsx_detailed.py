import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def extract_text_from_xlsx(xlsx_path, max_rows_per_sheet=10000):
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        result = []
        
        result.append(f"=== XLSX File: {os.path.basename(xlsx_path)} ===")
        result.append(f"Total Sheets: {len(wb.sheetnames)}")
        result.append(f"Sheet Names: {', '.join(wb.sheetnames)}")
        result.append("")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = min(ws.max_row, max_rows_per_sheet)
            max_col = ws.max_column
            
            result.append(f"=== Sheet: {sheet_name} ===")
            result.append(f"Dimensions: {max_row} rows x {max_col} columns")
            result.append("")
            
            # ヘッダー行を特定（最初の数行を確認）
            header_rows = min(5, max_row)
            result.append("--- Header Section (First 5 rows) ---")
            for row_idx in range(1, header_rows + 1):
                row_data = []
                for col_idx in range(1, min(max_col + 1, 50)):  # 最初の50列まで
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value).strip()
                        if cell_value:
                            row_data.append(cell_value)
                if row_data:
                    result.append(f"Row {row_idx}: {' | '.join(row_data)}")
            result.append("")
            
            # データ行を抽出（数値、日付、重要な情報を含む）
            result.append("--- Data Section (Sample rows with data) ---")
            sample_count = 0
            for row_idx in range(1, max_row + 1):
                row_data = []
                has_data = False
                
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = cell.value
                        
                        # 数値、日付、文字列を適切に処理
                        if isinstance(cell_value, (int, float)):
                            row_data.append(str(cell_value))
                            has_data = True
                        elif isinstance(cell_value, datetime):
                            row_data.append(cell_value.strftime('%Y-%m-%d %H:%M:%S'))
                            has_data = True
                        elif isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            if cell_value:
                                row_data.append(cell_value)
                                has_data = True
                        else:
                            cell_str = str(cell_value).strip()
                            if cell_str:
                                row_data.append(cell_str)
                                has_data = True
                
                if has_data and row_data:
                    result.append(f"Row {row_idx}: {' | '.join(row_data[:30])}")  # 最初の30列まで
                    sample_count += 1
                    if sample_count >= 500:  # 最大500行のサンプル
                        result.append(f"... (showing first 500 rows with data, total rows: {max_row})")
                        break
            
            result.append("")
        
        return "\n".join(result)
    except Exception as e:
        import traceback
        return f"Error extracting text from XLSX: {e}\n{traceback.format_exc()}"

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_xlsx_detailed.py <file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    result = extract_text_from_xlsx(file_path)

    # Ensure output is encoded in UTF-8 to handle special characters
    sys.stdout.buffer.write(result.encode('utf-8', errors='replace'))

