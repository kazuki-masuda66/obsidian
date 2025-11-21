import sys
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_text_from_xlsx(xlsx_path):
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        result = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append(f"=== Sheet: {sheet_name} ===")
            result.append(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}")
            result.append("")
            
            # 全てのセルからデータを抽出
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), start=1):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        # セルの値と型を取得
                        cell_value = cell.value
                        cell_type = type(cell_value).__name__
                        
                        # 数値の場合はそのまま、文字列の場合は文字列として
                        if isinstance(cell_value, (int, float)):
                            row_data.append(str(cell_value))
                        elif isinstance(cell_value, str):
                            row_data.append(cell_value)
                        else:
                            row_data.append(str(cell_value))
                
                if row_data:
                    result.append(f"Row {row_idx}: {' | '.join(row_data)}")
            
            result.append("")
        
        return "\n".join(result)
    except Exception as e:
        return f"Error extracting text from XLSX: {e}"

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_xlsx.py <file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    result = extract_text_from_xlsx(file_path)

    # Ensure output is encoded in UTF-8 to handle special characters
    sys.stdout.buffer.write(result.encode('utf-8', errors='replace'))

